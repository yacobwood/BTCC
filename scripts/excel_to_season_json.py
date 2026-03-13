#!/usr/bin/env python3
"""
Generate season_YYYY.json from excelData/Drivers.xlsx and Teams.xlsx.
Single source of truth: each JSON has drivers, teams, and rounds (race results).
Output: app/src/main/assets/data/season_2014.json ... season_2025.json

Reads from Excel only. Run after csv_to_standings.py (or standalone). Requires openpyxl.
"""

import json
import os
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

# Reuse parsing helpers from csv_to_standings
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from csv_to_standings import (
    repo_root,
    get_excel_paths,
    get_excel_years,
    get_sheet_name_for_year,
    read_sheet_rows,
    parse_drivers_from_rows,
    parse_teams_from_rows,
    _row_to_strings,
    _DRIVER_STOP_NAMES,
    DRIVER_LOOKUP,
    escape_kotlin_string,
)

def position_from_cell(val, cell=None, year=None):
    """Return (position_int, fastest_lap, lead_lap, pole).
    2021-2025: F = fastest lap, L = lead lap, P = R1 pole (letters in cell).
    2014-2020: Bold = pole, Italic = fastest lap, * = lead lap (formatting + * in value)."""
    if val is None:
        return 0, False, False, False
    v = str(val).strip()
    if not v:
        return 0, False, False, False
    u = v.upper()
    if u in ("RET", "DSQ", "DNS", "NC", "WD", "DNF", "C", "DNP", "DNA", "EX", "DNQ", "DNPQ"):
        return 0, False, False, False

    if year is not None and 2014 <= year <= 2020 and cell is not None:
        # 2014-2020: Bold = pole, Italic = fastest lap, * = lead lap
        try:
            font = cell.font
            pole = bool(font and getattr(font, "bold", False))
            fl = bool(font and getattr(font, "italic", False))
        except Exception:
            pole = fl = False
        lead = "*" in v
        # Position: strip * then parse digits
        pos_str = "".join(c for c in v if c.isdigit())
        try:
            pos = int(pos_str) if pos_str else 0
        except ValueError:
            pos = 0
        return pos, fl, lead, pole

    # 2021-2025: F, L, P letters (and * = F for legacy)
    digits = []
    rest = []
    for c in v:
        if c.isdigit():
            digits.append(c)
        else:
            rest.append(c.upper() if c != "*" else "F")
    pos_str = "".join(digits)
    suffix = "".join(rest)
    try:
        pos = int(pos_str) if pos_str else 0
    except ValueError:
        pos = 0
    fl = "F" in suffix or "*" in v
    lead = "L" in suffix
    pole = "P" in suffix
    return pos, fl, lead, pole

# BTCC position points for computing points per result
POINTS_BY_POS = [20, 17, 15, 13, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]

def points_from_position(pos):
    return POINTS_BY_POS[pos - 1] if 1 <= pos <= 15 else 0

# Expand Excel column-header abbreviations to full venue names (matches data/results20XX.json style)
VENUE_FULL_NAMES = {
    "BHI": "Brands Hatch Indy",
    "BHGP": "Brands Hatch GP",
    "DON": "Donington Park",
    "DONGP": "Donington Park GP",
    "THR": "Thruxton",
    "OUL": "Oulton Park",
    "CRO": "Croft",
    "SNE": "Snetterton",
    "KNO": "Knockhill",
    "ROC": "Rockingham",
    "SIL": "Silverstone",
}

def expand_venue(venue):
    if not venue:
        return venue
    key = venue.strip().upper()
    return VENUE_FULL_NAMES.get(key, venue)


def get_driver_sheet_cell_rows(wb, sheet_name, num_data_rows):
    """Return list of data rows as list of Cell objects (for 2014-2020 font reading)."""
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    # min_row=2: skip header; same row count as data_d
    rows = list(sheet.iter_rows(min_row=2, max_row=1 + num_data_rows, values_only=False))
    return [list(row) for row in rows]

def load_results_json_rounds(root, year):
    """Load full rounds array from data/results{year}.json if it exists (for dates and time data)."""
    path = os.path.join(root, "data", f"results{year}.json")
    if not os.path.isfile(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("rounds") or []
    except Exception:
        return None


def build_rounds_from_results_json(root, year):
    """
    Build rounds directly from data/results{year}.json with our BTCC scoring (single source of truth).
    Use when results JSON exists so every result gets correct pos/points without Excel merge issues.
    """
    raw_rounds = load_results_json_rounds(root, year)
    if not raw_rounds:
        return None
    out_rounds = []
    for r in raw_rounds:
        round_num = r.get("round", 0)
        venue = r.get("venue") or ""
        date = r.get("date") or ""
        races_out = []
        for race_idx, src_race in enumerate(r.get("races") or []):
            is_race1 = race_idx == 0
            results_out = []
            for res in src_race.get("results") or []:
                pos = res.get("pos", 0)
                fl = bool(res.get("fastestLap") or res.get("fl"))
                lead = bool(res.get("leadLap") or res.get("l"))
                pole = bool(res.get("pole") or res.get("p")) and is_race1
                # Optional override when source has pointsOverride (e.g. penalty applied in Excel)
                base_pts = res.get("pointsOverride")
                if isinstance(base_pts, (int, float)) and base_pts >= 0:
                    pts = base_pts + (1 if fl else 0) + (1 if lead else 0) + (1 if pole else 0)
                else:
                    pts = points_from_position(pos) + (1 if fl else 0) + (1 if lead else 0) + (1 if pole else 0)
                results_out.append({
                    "pos": pos,
                    "no": res.get("no", 0),
                    "driver": _title_case_driver((res.get("driver") or "").strip()),
                    "team": (res.get("team") or "").strip(),
                    "laps": res.get("laps", 0),
                    "time": res.get("time") or "",
                    "gap": res.get("gap") or "",
                    "bestLap": res.get("bestLap") or "",
                    "points": max(0, pts),
                    "fl": fl,
                    "l": lead,
                    "p": pole,
                })
            races_out.append({
                "label": src_race.get("label") or f"Race {race_idx + 1}",
                "date": (src_race.get("date") or "").strip() or None,
                "fullRaceUrl": (src_race.get("fullRaceUrl") or "").strip() or None,
                "results": results_out,
            })
        out_rounds.append({
            "round": round_num,
            "venue": venue,
            "date": date,
            "races": races_out,
        })
    return out_rounds

def load_round_dates_from_results_json(root, year):
    """Load round number -> date string from data/results{year}.json if it exists."""
    rounds_arr = load_results_json_rounds(root, year)
    if not rounds_arr:
        return {}
    return {r["round"]: (r.get("date") or "") for r in rounds_arr if "round" in r}

def _normalize_driver_name(name):
    return (name or "").strip().lower()


def _title_case_driver(name):
    """Normalise driver name for consistent casing (e.g. 'Jake HILL' -> 'Jake Hill')."""
    s = (name or "").strip()
    if not s:
        return s
    return " ".join(
        "-".join(p.capitalize() for p in w.split("-")) for w in s.split()
    )


# Excel/display name -> name as in results JSON (after title case). Used for points lookup.
DRIVER_NAME_ALIASES = {
    2014: {"Rob Collard": "Robert Collard", "Árón Smith": "Aron Taylor-Smith", "Daniel Welch": "Dan Welch"},
    2015: {"Rob Collard": "Robert Collard", "Árón Smith": "Aron Taylor-Smith", "Daniel Welch": "Dan Welch", "Derek Palmer Jr.": "Derek Palmer"},
    2016: {"Rob Collard": "Robert Collard", "Árón Smith": "Aron Taylor-Smith"},
    2017: {"Rob Collard": "Robert Collard", "Robert Huff": "Rob Huff", "Árón Taylor-Smith": "Aron Taylor-Smith"},
    2018: {"Rob Collard": "Robert Collard"},
    2019: {"Rob Collard": "Robert Collard"},
    2020: {"Nicolas Hamilton": "Nic Hamilton"},
    2021: {"Árón Taylor-Smith": "Aron Taylor-Smith"},
    2022: {"Árón Taylor-Smith": "Aron Taylor-Smith"},
    2023: {"Árón Taylor-Smith": "Aron Taylor-Smith", "Nicolas Hamilton": "Nic Hamilton", "Robert Huff": "Rob Huff", "Daryl DeLeon": "Daryl Deleon"},
    2024: {"Daryl DeLeon": "Daryl Deleon"},
    2025: {"Daryl DeLeon": "Daryl Deleon"},
}

def merge_times_and_teams_from_results_json(root, year, rounds, merge_positions=False):
    """
    Merge time, gap, bestLap, team, laps (and race date/fullRaceUrl) from data/results{year}.json
    into the Excel-built rounds. Matches by round number, race index, and driver name.
    When merge_positions=False (default): only merge display data; pos/points stay from Excel.
    When merge_positions=True: also overwrite pos and points from the results JSON (legacy).
    """
    results_rounds = load_results_json_rounds(root, year)
    if not results_rounds:
        return
    by_round = {r["round"]: r for r in results_rounds if "round" in r}
    for round_obj in rounds:
        round_num = round_obj.get("round")
        src = by_round.get(round_num)
        if not src:
            continue
        src_races = src.get("races") or []
        for race_idx, race in enumerate(round_obj.get("races") or []):
            if race_idx >= len(src_races):
                break
            src_race = src_races[race_idx]
            if not race.get("date") and src_race.get("date"):
                race["date"] = src_race["date"]
            if src_race.get("fullRaceUrl"):
                race["fullRaceUrl"] = src_race["fullRaceUrl"]
            src_results = src_race.get("results") or []
            by_driver = {_normalize_driver_name(r.get("driver")): r for r in src_results}
            is_race1 = race_idx == 0
            for res in race.get("results") or []:
                key = _normalize_driver_name(res.get("driver"))
                src_res = by_driver.get(key)
                if not src_res:
                    continue
                for field in ("time", "gap", "bestLap", "team", "laps"):
                    val = src_res.get(field)
                    if val is not None and (val != "" or field == "team"):
                        res[field] = val
                if "no" in src_res and src_res.get("no"):
                    res["no"] = src_res["no"]
                if merge_positions:
                    pos = src_res.get("pos", 0)
                    fl = bool(src_res.get("fastestLap") or src_res.get("fl"))
                    lead = bool(src_res.get("leadLap") or src_res.get("l"))
                    pole = bool(src_res.get("pole") or src_res.get("p")) and is_race1
                    pts = points_from_position(pos) + (1 if fl else 0) + (1 if lead else 0) + (1 if pole else 0)
                    res["pos"] = pos
                    res["points"] = max(0, pts)
                    res["fl"] = fl
                    res["l"] = lead
                    res["p"] = pole

def driver_totals_from_rounds(rounds):
    """Sum points per driver from rounds (single source of truth for chart and standings)."""
    totals = {}
    for round_obj in rounds:
        for race in round_obj.get("races") or []:
            for res in race.get("results") or []:
                name = (res.get("driver") or "").strip()
                if name:
                    totals[name] = totals.get(name, 0) + res.get("points", 0)
    return totals


def compute_driver_stats(rounds):
    """Precompute wins, podiums, poles, dnfs, races per driver (Stats tab – no app calculation)."""
    # driver -> { team, races, wins, podiums, poles, dnfs }
    acc = {}
    for round_obj in sorted(rounds, key=lambda r: r.get("round", 0)):
        pole_driver = (round_obj.get("polePosition") or "").strip()
        if pole_driver:
            acc.setdefault(pole_driver, {"team": "", "races": 0, "wins": 0, "podiums": 0, "poles": 0, "dnfs": 0})["poles"] += 1
        for race in round_obj.get("races") or []:
            for res in race.get("results") or []:
                name = (res.get("driver") or "").strip()
                if not name:
                    continue
                r = acc.setdefault(name, {"team": "", "races": 0, "wins": 0, "podiums": 0, "poles": 0, "dnfs": 0})
                if not r["team"]:
                    r["team"] = (res.get("team") or "").strip()
                r["races"] += 1
                pos = res.get("pos", 0)
                if pos == 1:
                    r["wins"] += 1
                    r["podiums"] += 1
                elif pos in (2, 3):
                    r["podiums"] += 1
                elif pos <= 0:
                    r["dnfs"] += 1
    out = []
    for driver, r in acc.items():
        out.append({
            "driver": driver,
            "team": r["team"],
            "races": r["races"],
            "wins": r["wins"],
            "podiums": r["podiums"],
            "poles": r["poles"],
            "dnfs": r["dnfs"],
        })
    out.sort(key=lambda x: (-x["wins"], -x["podiums"], -x["poles"], -x["races"]))
    return out


def compute_progression(rounds):
    """Precompute cumulative points by round per driver (Chart tab – no app calculation)."""
    sorted_rounds = sorted(rounds, key=lambda r: r.get("round", 0))
    points_per_round = {}  # driver -> [pts_r1, pts_r2, ...]
    team_by_driver = {}
    for round_obj in sorted_rounds:
        for race in round_obj.get("races") or []:
            for res in race.get("results") or []:
                name = (res.get("driver") or "").strip()
                if not name:
                    continue
                if name not in points_per_round:
                    points_per_round[name] = []
                    team_by_driver[name] = (res.get("team") or "").strip()
    # per-round points (same length for all: one entry per round)
    for round_obj in sorted_rounds:
        round_pts = {}
        for race in round_obj.get("races") or []:
            for res in race.get("results") or []:
                name = (res.get("driver") or "").strip()
                if name:
                    round_pts[name] = round_pts.get(name, 0) + res.get("points", 0)
        for name in points_per_round:
            points_per_round[name].append(round_pts.get(name, 0))
    # cumulative
    cumulative = {}
    for driver, per_round in points_per_round.items():
        cum = []
        s = 0
        for p in per_round:
            s += p
            cum.append(s)
        cumulative[driver] = cum
    out = []
    for driver in sorted(cumulative.keys(), key=lambda d: -(cumulative[d][-1] if cumulative[d] else 0)):
        out.append({
            "driver": driver,
            "team": team_by_driver.get(driver, ""),
            "cumulativePointsByRound": cumulative[driver],
        })
    return out


def _format_gap(gap):
    """Format gap from leader for display: '+X.XXX' or None."""
    if gap is None or (isinstance(gap, str) and not gap.strip()):
        return None
    s = str(gap).strip()
    return s if s.startswith("+") else f"+{s}"

def add_display_times_to_rounds(rounds):
    """
    Set displayTime on each result so the app can show it without calculation.
    P1: leader's full time (time or bestLap). P2+: '+X.XXX' (gap from leader).
    """
    for round_obj in rounds:
        for race in round_obj.get("races") or []:
            results = race.get("results") or []
            if not results:
                continue
            first = results[0]
            leader_display = (first.get("time") or "").strip() or (first.get("bestLap") or "").strip() or "—"
            for i, res in enumerate(results):
                if i == 0:
                    res["displayTime"] = leader_display
                else:
                    res["displayTime"] = _format_gap(res.get("gap")) or "—"

def build_rounds_from_driver_sheet(header_row, data_rows, round_dates=None, cell_rows=None, year=None):
    """Build rounds[] from header and data rows. cell_rows + year used for 2014-2020 (Bold/Italic/*)."""
    header = _row_to_strings(header_row)
    n_cols = len(header)
    pts_col = n_cols - 1
    n_race_cols = pts_col - 2
    if n_race_cols < 3 or n_race_cols % 3 != 0:
        return []
    n_rounds = n_race_cols // 3
    rounds = []
    for r in range(n_rounds):
        raw_venue = (header[2 + r * 3] or "").strip() or f"Round {r+1}"
        venue = expand_venue(raw_venue) if len(raw_venue) <= 6 else raw_venue
        races = []
        for race_idx in range(3):
            col = 2 + r * 3 + race_idx
            results = []
            for row_idx, row in enumerate(data_rows):
                row = _row_to_strings(row)
                if len(row) <= col:
                    continue
                name = (row[1] or "").strip()
                if not name or name.lower() in _DRIVER_STOP_NAMES:
                    break
                pos_val = row[col] if col < len(row) else ""
                cell = None
                if cell_rows and row_idx < len(cell_rows) and col < len(cell_rows[row_idx]):
                    cell = cell_rows[row_idx][col]
                pos, fl, lead, pole = position_from_cell(pos_val, cell=cell, year=year)
                is_race1 = (race_idx == 0)
                pts = points_from_position(pos) + (1 if fl else 0) + (1 if lead else 0) + (1 if (pole and is_race1) else 0)
                results.append({
                    "pos": pos,
                    "no": 0,
                    "driver": name,
                    "team": "",
                    "laps": 0,
                    "time": "",
                    "gap": "",
                    "bestLap": "",
                    "points": pts,
                    "fl": fl,
                    "l": lead,
                    "p": pole and is_race1,
                })
            # Sort: position 1,2,3... first, then 0 (Ret etc)
            results.sort(key=lambda x: (x["pos"] == 0, x["pos"]))
            for i, res in enumerate(results):
                res["no"] = i + 1
            races.append({"label": f"Race {race_idx + 1}", "results": results})
        round_num = r + 1
        date = (round_dates or {}).get(round_num, "")
        rounds.append({
            "round": round_num,
            "venue": venue,
            "date": date,
            "races": races,
        })
    return rounds

def driver_standings_to_json(drivers, year):
    """Convert (pos, name, pts, wins) list to JSON driver standings with team/car from lookup."""
    lookup = DRIVER_LOOKUP.get(year, {})
    out = []
    for pos, name, pts, wins in drivers:
        team, car, team_sec = lookup.get(name, ("", "", None))
        team_sec = team_sec or ""
        o = {"position": pos, "name": name, "team": team, "points": pts, "wins": wins}
        if car:
            o["car"] = car
        if team_sec:
            o["teamSecondary"] = team_sec
        out.append(o)
    return out

def team_standings_to_json(teams):
    return [{"position": pos, "name": team, "points": pts} for pos, team, pts in teams]

def main():
    root = repo_root()
    drivers_path, teams_path = get_excel_paths(root)
    if not drivers_path or not teams_path:
        print("Put Drivers.xlsx and Teams.xlsx in excelData/")
        return 1
    years = get_excel_years(root)
    if not years:
        print("No year sheets found")
        return 1

    out_dir = os.path.join(root, "app", "src", "main", "assets", "data")
    os.makedirs(out_dir, exist_ok=True)

    # Drivers: data_only=True for correct cell values (2021+ F/L/P and numbers). Second workbook for font (2014-2020).
    wb_d_values = load_workbook(drivers_path, read_only=True, data_only=True)
    wb_d_cells = load_workbook(drivers_path, read_only=False, data_only=False)
    wb_t = load_workbook(teams_path, read_only=True, data_only=True)

    for year in years:
        sheet_d = get_sheet_name_for_year(wb_d_values, year)
        sheet_t = get_sheet_name_for_year(wb_t, year)
        if not sheet_d or not sheet_t:
            continue
        header_d, data_d = read_sheet_rows(wb_d_values, sheet_d)
        header_t, data_t = read_sheet_rows(wb_t, sheet_t)
        if not header_d or not data_d:
            continue
        if not header_t or not data_t:
            continue

        drivers = parse_drivers_from_rows(data_d, header_d)
        teams = parse_teams_from_rows(data_t, header_t)
        if not drivers:
            continue

        # Always build rounds from Excel (single source of truth for positions and points)
        round_dates = load_round_dates_from_results_json(root, year)
        cell_rows = get_driver_sheet_cell_rows(wb_d_cells, sheet_d, len(data_d)) if 2014 <= year <= 2020 else None
        rounds = build_rounds_from_driver_sheet(header_d, data_d, round_dates, cell_rows=cell_rows, year=year)
        # Optionally enrich with time/gap/team from results JSON (no pos/points overwrite)
        merge_times_and_teams_from_results_json(root, year, rounds, merge_positions=False)
        add_display_times_to_rounds(rounds)

        # Single source of truth: driver totals from rounds (chart and Drivers tab use same figures)
        points_from_rounds = driver_totals_from_rounds(rounds)
        # Add alias lookups so Excel names (e.g. Rob Collard) match round names (Robert Collard)
        aliases = DRIVER_NAME_ALIASES.get(year, {})
        for excel_name, round_name in aliases.items():
            if round_name in points_from_rounds and excel_name not in points_from_rounds:
                points_from_rounds[excel_name] = points_from_rounds[round_name]
        drivers = [(pos, name, points_from_rounds.get(name, points_from_rounds.get(_title_case_driver(name), pts)), wins) for pos, name, pts, wins in drivers]

        # Precompute stats and progression so the app does zero calculation (hardcoded data only)
        driver_stats = compute_driver_stats(rounds)
        progression = compute_progression(rounds)

        payload = {
            "season": str(year),
            "drivers": driver_standings_to_json(drivers, year),
            "teams": team_standings_to_json(teams),
            "rounds": rounds,
            "driverStats": driver_stats,
            "progression": progression,
        }
        path = os.path.join(out_dir, f"season_{year}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        print(f"Wrote {path} ({len(drivers)} drivers, {len(teams)} teams, {len(rounds)} rounds)", flush=True)

    wb_d_values.close()
    wb_d_cells.close()
    wb_t.close()
    return 0

if __name__ == "__main__":
    sys.exit(main() or 0)
