#!/usr/bin/env python3
"""
Generate Kotlin Standings from Drivers and Teams data.

Input: Drivers.xlsx and Teams.xlsx in excelData/ — each sheet/tab = one year
       (sheet names like "2014", "2015", or "Season 2014").

Output: app/src/main/java/com/btccfanhub/data/Standings{year}.kt for each year.

Install: pip install openpyxl
"""

import os
import re
import sys

try:
    from openpyxl import load_workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# Optional: driver name -> (team, car, team_secondary) per year. If missing, uses "".
DRIVER_LOOKUP = {
    2014: {
        "Colin Turkington": ("eBay Motors", "5", None),
        "Jason Plato": ("MG KX Clubcard Fuel Save", "99", None),
        "Gordon Shedden": ("Honda Yuasa Racing", "52", None),
        "Mat Jackson": ("Airwaves Racing", "6", None),
        "Andrew Jordan": ("Pirtek Racing", "1", None),
        "Rob Collard": ("eBay Motors", "10", None),
        "Sam Tordoff": ("MG KX Clubcard Fuel Save", "88", None),
        "Matt Neal": ("Honda Yuasa Racing", "4", None),
        "Árón Smith": ("CHROME Edition Restart Racing", "40", None),
        "Adam Morgan": ("WIX Racing", "33", None),
        "Alain Menu": ("CHROME Edition Restart Racing", "9", None),
        "Rob Austin": ("Exocet Racing", "101", None),
        "Fabrizio Giovanardi": ("Airwaves Racing", "7", None),
        "Tom Ingram": ("Speedworks Motorsport", "80", None),
        "Jack Goff": ("CHROME Edition Restart Racing", "31", "RCIB Insurance Racing"),
        "Nick Foster": ("eBay Motors", "18", None),
        "Dave Newsham": ("AmD Tuning.com", "17", None),
        "Marc Hynes": ("Quantel BiFold Racing", "888", None),
        "Jack Clarke": ("Crabbie's Racing", "44", None),
        "Hunter Abbott": ("AlcoSense Breathalysers Racing", "54", None),
        "Warren Scott": ("CHROME Edition Restart Racing", "39", None),
        "Glynn Geddie": ("United Autosports", "21", None),
        "Martin Depper": ("Pirtek Racing", "30", None),
        "Lea Wood": ("Houseman Racing", "43", None),
        "Aiden Moffat": ("Laser Tools Racing", "16", None),
        "James Cole": ("United Autosports", "20", None),
        "Robb Holland": ("Rotek Racing", "67", None),
        "Ollie Jackson": ("STP Racing with Sopp & Sopp", "48", None),
        "Luke Hines": ("United Autosports", "23", None),
        "Simon Belcher": ("Handy Motorsport", "11", None),
        "Chris Stockton": ("Power Maxed Racing", "28", None),
        "Daniel Welch": ("STP Racing with Sopp & Sopp", "12", None),
    },
    2020: {
        "Jade Edwards": ("BTC Racing", "99", None),
        "Jac Constable": ("Power Maxed Car Care Racing", "96", None),
    },
    2021: {
        "Jade Edwards": ("BTC Racing", "99", None),
    },
    2022: {
        "Jade Edwards": ("Rich Energy BTC Racing", "99", None),
        "Carl Boardley": ("Rich Energy BTC Racing", "29", None),
    },
}

def repo_root():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def parse_int(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s) if s == int(s) else None
    s = str(s).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None

def is_race_result_cell(val):
    if val is None:
        return False
    v = str(val).strip()
    if not v:
        return False
    if re.match(r"^-?\d+", v):
        return True
    if v.upper() in ("RET", "DSQ", "DNS", "NC", "WD", "DNF", "C", "DNP", "DNA", "EX", "DNQ", "DNPQ"):
        return True
    return False

def count_wins(row, header):
    wins = 0
    for i, h in enumerate(header):
        if i < 2 or (h and str(h).strip() == "Pts"):
            continue
        v = (row[i] if i < len(row) else "")
        if not is_race_result_cell(v):
            continue
        v = str(v).strip()
        if v.startswith("1") and (v == "1" or v == "1*"):
            wins += 1
    return wins

_DRIVER_STOP_NAMES = frozenset({
    "driver", "colour", "result", "pos", "gold", "silver", "bronze", "green", "blue",
    "purple", "red", "black", "white", "blank", "winner", "second place", "third place",
    "points classification", "non-points classification", "retired, not classified (ret)",
    "did not qualify (dnq)", "did not pre-qualify (dnpq)", "disqualified (dsq)",
    "did not start (dns)", "withdrew (wd)", "race cancelled (c)", "did not practice (dnp)",
    "did not arrive (dna)", "excluded (ex)", "non-classified finish (nc)",
})

def _row_to_strings(row):
    out = []
    for c in row:
        if c is None:
            out.append("")
        elif isinstance(c, float) and c == int(c):
            out.append(str(int(c)))
        else:
            out.append(str(c).strip())
    return out

def parse_drivers_from_rows(rows, header_row):
    """Parse driver list from rows (list of lists). header_row is first row; rest are data."""
    drivers = []
    header = _row_to_strings(header_row)
    for row in rows:
        row = _row_to_strings(row)
        if len(row) < 3:
            continue
        pos = parse_int(row[0])
        name = (row[1] or "").strip()
        if not name:
            continue
        if name.lower() in _DRIVER_STOP_NAMES:
            break
        if pos is None:
            continue
        pts_val = row[-1].strip() if row else ""
        try:
            pts = int(pts_val) if pts_val else 0
        except ValueError:
            pts = 0
        wins = count_wins(row, header)
        drivers.append((pos, name, pts, wins))
    return drivers

def parse_teams_from_rows(rows, header_row):
    teams = []
    for row in rows:
        row = _row_to_strings(row)
        if len(row) < 3:
            continue
        pos = parse_int(row[0])
        team = (row[1] or "").strip()
        if not team:
            continue
        if team.lower() in ("team", "colour", "result"):
            break
        if pos is None:
            continue
        pts_val = row[-1].strip() if row else ""
        try:
            pts = int(pts_val) if pts_val else 0
        except ValueError:
            pts = 0
        teams.append((pos, team, pts))
    return teams

def extract_year_from_sheet_name(name):
    """Return 4-digit year from sheet name, e.g. '2014', 'Season 2014', '14' -> 2014."""
    name = str(name).strip()
    # Exact 4-digit year
    m = re.search(r"\b(201[4-9]|202[0-9])\b", name)
    if m:
        return int(m.group(1))
    # Two-digit year
    m = re.search(r"\b(1[4-9]|2[0-9])\b", name)
    if m:
        y = int(m.group(1))
        return 2000 + y if y < 50 else 1900 + y
    return None

def read_sheet_rows(wb, sheet_name):
    """Read one sheet from an open workbook. Returns (header_row, data_rows) or (None, None)."""
    if sheet_name not in wb.sheetnames:
        return None, None
    sheet = wb[sheet_name]
    rows = []
    for r in sheet.iter_rows(values_only=True):
        rows.append(list(r))
    if not rows:
        return None, None
    return rows[0], rows[1:]

def get_excel_paths(root):
    """Return (drivers_path, teams_path) for Excel files in excelData/ or root."""
    for folder in (os.path.join(root, "excelData"), root):
        d = os.path.join(folder, "Drivers.xlsx")
        t = os.path.join(folder, "Teams.xlsx")
        if os.path.isfile(d) and os.path.isfile(t):
            return d, t
    return None, None

def get_excel_years(root):
    """Return sorted list of years that exist in both Drivers.xlsx and Teams.xlsx."""
    drivers_path, teams_path = get_excel_paths(root)
    if not drivers_path or not teams_path:
        return []
    if not os.path.isfile(drivers_path) or not os.path.isfile(teams_path):
        return []
    wb_d = load_workbook(drivers_path, read_only=True)
    wb_t = load_workbook(teams_path, read_only=True)
    years_d = set()
    for name in wb_d.sheetnames:
        y = extract_year_from_sheet_name(name)
        if y and 2014 <= y <= 2030:
            years_d.add(y)
    years_t = set()
    for name in wb_t.sheetnames:
        y = extract_year_from_sheet_name(name)
        if y and 2014 <= y <= 2030:
            years_t.add(y)
    wb_d.close()
    wb_t.close()
    return sorted(years_d & years_t)

def get_sheet_name_for_year(wb, year):
    """Find sheet name that corresponds to year (exact '2014' or contains year)."""
    for name in wb.sheetnames:
        if extract_year_from_sheet_name(name) == year:
            return name
    return None

def escape_kotlin_string(s):
    if not s:
        return s
    return (
        str(s)
        .replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\r\n", " ")
        .replace("\n", " ")
        .replace("\r", " ")
    )

def emit_driver(driver_lookup, pos, name, pts, wins):
    team, car, team_sec = driver_lookup.get(name, ("", "", None))
    team_sec = team_sec or ""
    args = [
        f"position = {pos:2d}",
        f'name = "{escape_kotlin_string(name)}"',
        f'team = "{escape_kotlin_string(team)}"',
    ]
    if team_sec:
        args.append(f'teamSecondary = "{escape_kotlin_string(team_sec)}"')
    args.append(f'car = "{car}"' if car else 'car = ""')
    args.append(f"points = {pts}")
    args.append(f"wins = {wins}")
    return "DriverStanding(" + ", ".join(args) + ")"

def write_standings_kt(root, year, drivers, teams):
    out_path = os.path.join(root, "app", "src", "main", "java", "com", "btccfanhub", "data", f"Standings{year}.kt")
    lookup = DRIVER_LOOKUP.get(year, {})
    lines = [
        "package com.btccfanhub.data",
        "",
        "import com.btccfanhub.data.model.DriverStanding",
        "import com.btccfanhub.data.model.TeamStanding",
        "",
        f"/** Final {year} BTCC standings from Drivers and Teams (Excel). */",
        f"object Standings{year} {{",
        "",
        "    val drivers: List<DriverStanding> = listOf(",
    ]
    for pos, name, pts, wins in drivers:
        lines.append("        " + emit_driver(lookup, pos, name, pts, wins) + ",")
    lines.append("    )")
    lines.append("")
    lines.append("    val teams: List<TeamStanding> = listOf(")
    for pos, team, pts in teams:
        lines.append(f'        TeamStanding(position = {pos}, name = "{escape_kotlin_string(team)}", points = {pts}),')
    lines.append("    )")
    lines.append("}")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return out_path

def run_from_excel(root):
    drivers_path, teams_path = get_excel_paths(root)
    if not drivers_path or not teams_path:
        return False
    years = get_excel_years(root)
    if not years:
        return False
    wb_d = load_workbook(drivers_path, read_only=True, data_only=True)
    wb_t = load_workbook(teams_path, read_only=True, data_only=True)
    for year in years:
        sheet_d = get_sheet_name_for_year(wb_d, year)
        sheet_t = get_sheet_name_for_year(wb_t, year)
        if not sheet_d or not sheet_t:
            continue
        header_d, data_d = read_sheet_rows(wb_d, sheet_d)
        header_t, data_t = read_sheet_rows(wb_t, sheet_t)
        if not header_d or not data_d:
            continue
        if not header_t or not data_t:
            continue
        drivers = parse_drivers_from_rows(data_d, header_d)
        teams = parse_teams_from_rows(data_t, header_t)
        if not drivers:
            continue
        out = write_standings_kt(root, year, drivers, teams)
        print(f"Wrote {out} ({len(drivers)} drivers, {len(teams)} teams)", flush=True)
    wb_d.close()
    wb_t.close()
    return True

def main():
    root = repo_root()
    drivers_xlsx, teams_xlsx = get_excel_paths(root)

    if not drivers_xlsx or not teams_xlsx:
        print("Put Drivers.xlsx and Teams.xlsx in excelData/ (each sheet = one year).")
        print("Then run: python3 csv_to_standings.py")
        sys.exit(1)
    if not HAS_EXCEL:
        print("openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)
    if not run_from_excel(root):
        print("No valid year sheets found in Drivers.xlsx / Teams.xlsx (use sheet names like 2014, 2015)")
        sys.exit(1)

if __name__ == "__main__":
    main()
